# -*- coding: utf-8 -*
import os
import gc
import re
from argparse import ArgumentParser
from glob import glob
from math import ceil
from collections import OrderedDict
import numpy as np
import openpyxl
import seaborn as sns
import torch
import torch.nn as nn
import torch.nn.functional as F
import torchvision
from matplotlib import pyplot as plt
from pandas import DataFrame

from torch.utils.data import DataLoader
from prefetch_generator import BackgroundGenerator

from Utils import utils
from EBV_classifier.Dataloader_test import Dataset_test

from Model.create_model import create_model

plt.switch_backend('agg')

class DataLoaderX(DataLoader):
    def __iter__(self):
        return BackgroundGenerator(super().__iter__())



def get_index(path):
    path = os.path.basename(path)
    y, x = list(map(int, re.findall(r'\d+', path)[-3: -1]))
    return x, y


def load_xlsx_file(xlsxFile):
    label_dict = {}
    wb = openpyxl.load_workbook(xlsxFile)
    ws = wb.active
    max_row = ws.max_row

    for row in range(2, max_row + 1):
        label_dict[ws.cell(row, 2).value] = ws.cell(row, 3).value
    return label_dict

Map = {
    'EBV': 1,
    'ELSE': 0,
}

if __name__ == '__main__':

    parser = ArgumentParser()
    parser.add_argument("--Model1Path", dest='model1_path', type=str, help='the pretrained model of tumor detector')
    parser.add_argument("--Model2Path", dest='model2_path', type=str, help='the pretrained model of EBV classifier')
    parser.add_argument("--DataPath", dest='data_path',
                        default='./data/TCGA/Vahadane_TCGA_patch_10x_all', type=str, help='the path of the data need to test')
    parser.add_argument("--DeviceId", dest='device_id', type=str, help='select the gpu id to use')
    parser.add_argument("--Seed", dest='seed', default=0, type=str, help='the random seed')
    parser.add_argument("--Model1", dest='model1', default='resnet50', type=str, help='the model of the trained model for tumor detector')
    parser.add_argument("--Model2", dest='model2', default='resnet18', type=str, help='the model of the trained model for EBV classifier')
    parser.add_argument("--Labelpath", dest='labelpath',
                        default='./data/TCGA/TCGA_label.xlsx', type=str, help='the file content the slides name and their type')
    parser.add_argument("--SavePath", dest='save_path',
                        default='seed_result', type=str, help='the path to save the test result')
    parser.add_argument("--Fold", dest='fold', type=int, help='')
    parser.add_argument("--BatchSize", dest='batch_size', default=32, type=int)
    parser.add_argument("--Data_type", dest='data_tyoe', default='TCGA', type=str, help='the type of the test data: TCGA multicenter inner')
    args = parser.parse_args()

    os.environ['CUDA_VISIBLE_DEVICES'] = args.device_id

    if args.device_id is not None:
        device = torch.device('cuda:0' if torch.cuda.is_available() else 'cpu')

    state_dict1 = torch.load(args.model1_path)
    new_state_dict1 = OrderedDict()
    for k, v in state_dict1.items():
        new_state_dict1[k] = v
    model1 = create_model(args.model1, False)
    for key in state_dict1.keys():
        if 'module.' in key:
            print('parallel')
            model1 = torch.nn.DataParallel(model1)
            break
    model1_dict = model1.state_dict()
    weights1 = {k: v for k, v in new_state_dict1.items() if k in model1_dict}
    model1_dict.update(weights1)
    model1.load_state_dict(model1_dict, True)
    model1 = model1.to(device)

    state_dict2 = torch.load(args.model2_path)
    model2 = create_model(args.model2, False)
    for key in state_dict2.keys():
        if 'module.' in key:
            print('parallel')
            model2 = torch.nn.DataParallel(model2)
            break
    model2.load_state_dict(state_dict2, True)
    model2 = model2.to(device)

    if torch.cuda.device_count() == 2:
        print(os.environ['CUDA_VISIBLE_DEVICES'])
        model1 = torch.nn.DataParallel(model1, device_ids=[0, 1])
        model2 = torch.nn.DataParallel(model2, device_ids=[0, 1])
    elif torch.cuda.device_count() == 4:
        print(os.environ['CUDA_VISIBLE_DEVICES'])
        model1 = torch.nn.DataParallel(model1, device_ids=[0, 1, 2, 3])
        model2 = torch.nn.DataParallel(model2, device_ids=[0, 1, 2, 3])

    model1.eval()
    model2.eval()

    path = args.data_path
    test_path = path
    if args.data_type == 'inner':
        EBV_cases = glob(os.path.join(test_path, 'EBV', '*'))
        ELSE_cases = glob(os.path.join(test_path, 'ELSE', '*'))
        test_cases = EBV_cases + ELSE_cases
    else:
        test_cases = glob(os.path.join(test_path, '*'))
        label_dict = load_xlsx_file(args.labelpath, args.data_type)

    slides, ground_truth, aver, counts = [], [], [], []
    all_labels, all_predicts, all_values = [], [], []



    cnt = 0

    for case_name in test_cases:
        gc.collect()
        cnt += 1
        slides.append(os.path.basename(case_name))

        groundtruth = label_dict[os.path.basename(case_name)]
        if not groundtruth == 'EBV':
            groundtruth = 'ELSE'

        ground_truth.append(groundtruth)

        # FIXME: if args.data_type == 'inner' , sample_path = case_name
        sample_path = os.path.join(case_name, '10.0')
        print('sample_path:', sample_path)
        test_dataset = Dataset_test(path=sample_path, img_type=groundtruth, transforms=utils.transform_test, padding=0,
                                    bi=True)
        sample_num = test_dataset.__len__()
        print('sample_num', sample_num)

        if sample_num == 0:
            aver.append([-1, -1])
            counts.append([-1, -1])
            continue

        current_patch = 0
        im = []
        cases = []
        per = []
        test_loader = DataLoader(test_dataset, batch_size=args.batch_size, num_workers=4, pin_memory=False, persistent_workers=True)
        rlt_aver1 = np.zeros([2, sample_num])
        rlt_count1 = np.zeros([2, sample_num])
        for (inputs, labels) in test_loader:
            tmp = len(labels)
            input = inputs.to(device)
            output = model1(input)
            output = F.softmax(output, dim=1)
            # print(output)
            value, pred = torch.max(output, 1)
            output, v, p = output.detach().cpu().numpy(), value.detach().cpu().numpy(), pred.detach().cpu().numpy()
            for j in range(tmp):
                for index in range(2):
                    rlt_aver1[index][current_patch + j] = output[j][index]
                    rlt_count1[index][current_patch + j] = (pred[j] == index)
                if pred[j] == 1:
                    cases.append(inputs[j])
            current_patch += tmp
        average1 = rlt_aver1.mean(axis=1)  # 预测值的平均
        count1 = rlt_count1.mean(axis=1)  # 阴性阳性预测数量的平均
        print('tumor patch: ', len(cases))
        print(average1, count1)
        if len(cases) == 0:
            slides = slides[:-1]
            ground_truth = ground_truth[:-1]
            continue

        current_patch = 0
        rlt_aver = np.zeros([2, len(cases)])
        rlt_count = np.zeros([2, len(cases)])
        bs = args.batch_size
        for i in range(ceil(len(cases) / bs)):
            im = []
            tmp = 0
            for j in range(bs):
                if tmp + current_patch < len(cases):
                    pic = cases[tmp + current_patch]
                    im.append(pic)
                    tmp += 1
            input = torch.stack(im, dim=0)
            input = input.to(device)
            output = model2(input)
            output = F.softmax(output, dim=1)
            value, pred = torch.max(output, 1)
            output, v, p = output.detach().cpu().numpy(), value.detach().cpu().numpy(), pred.detach().cpu().numpy()
            for j in range(tmp):
                for index in range(2):
                    rlt_aver[index][current_patch + j] = output[j][index]
                    rlt_count[index][current_patch + j] = (pred[j] == index)
            current_patch += tmp
        average = rlt_aver.mean(axis=1)  # 预测值的平均
        count = rlt_count.mean(axis=1)  # 阴性阳性预测数量的平均

        print(average, count)
        aver.append(average)
        counts.append(count)
        all_labels.append(Map[groundtruth])
        all_predicts.append(np.argmax(average))
        all_values.append(average[1])
        print(average[1])
        print("label: ", Map[groundtruth])
        print("predict: ", np.argmax(average))
        print("model name:{}  fold:{}  speed:{}/{}".format(args.model2, args.fold, cnt, len(test_cases)))

    all_labels = np.array(all_labels)
    all_predicts = np.array(all_predicts)
    all_values = np.array(all_values)
    TP = ((all_predicts == 1) & (all_labels == 1)).sum()
    TN = ((all_predicts == 0) & (all_labels == 0)).sum()
    FN = ((all_predicts == 0) & (all_labels == 1)).sum()
    FP = ((all_predicts == 1) & (all_labels == 0)).sum()
    print('TP: %d, TN:%d, FN:%d, FP:%d' % (TP, TN, FN, FP))

    p = TP / (TP + FP)
    r = TP / (TP + FN)
    F1 = 2 * r * p / (r + p)
    print('EBV: precision: {:.4f}, recall: {:.4f}, F1: {:.4f}'.format(p, r, F1))
    print("Acc: ", (all_labels == all_predicts).sum() / len(all_predicts))


    df = DataFrame(
        {'Slide:': slides,
         'Type:': ground_truth,
         'Averaging': aver,
         'Counting:': counts}
    )
    save_path = os.path.join(args.save_path, args.data_type + '_' + 'seed' + str(args.seed) + '_' 'fold' + str(args.fold) + '.csv')
    df.to_csv(save_path)


