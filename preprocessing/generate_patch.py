from __future__ import print_function
import json
import openslide
from openslide import open_slide, ImageSlide
from openslide.deepzoom import DeepZoomGenerator
from optparse import OptionParser
import re
import shutil
from unicodedata import normalize
import numpy as np
import scipy.misc
import subprocess
from glob import glob
from multiprocessing import Process, JoinableQueue
import time
import os
import sys
import dicom
from collections import Counter
import openpyxl

from xml.dom import minidom
from PIL import Image, ImageDraw, ImageCms
from skimage import color, io
Image.MAX_IMAGE_PIXELS = None
VIEWER_SLIDE_NAME = 'slide'
tumor_threshold = 0.5
label_downsample = 8



def load_xlsx_file(xlsxFile):
    label_dict = {}
    wb = openpyxl.load_workbook(xlsxFile)
    ws = wb.active
    max_row = ws.max_row
    
    for row in range(2, max_row + 1):
        label_dict[ws.cell(row, 1).value] = ws.cell(row, 2).value
        
    return label_dict

class TileWorker(Process):
    """A child process that generates and writes tiles."""

    def __init__(self, queue, slidepath, tile_size, overlap, limit_bounds, quality, _Bkg):
        Process.__init__(self, name='TileWorker')
        self.daemon = True
        self._queue = queue
        self._slidepath = slidepath
        self._tile_size = tile_size
        self._overlap = overlap
        self._limit_bounds = limit_bounds
        self._quality = quality
        self._slide = None
        self._Bkg = _Bkg

    def run(self):
        self._slide = open_slide(self._slidepath)
        dz = self._get_dz()
        while True:
            data = self._queue.get()
            if data is None:
                self._queue.task_done()
                break
            # associated, level, address, outfile = data
            level, address, outfile = data

            try:
                tile = dz.get_tile(level, address)
                #gray = tile.convert('L')
                #bw = gray.point(lambda x: 0 if x < 220 else 1, 'F')
                #avgBkg = np.average(bw)
                #bw = gray.point(lambda x: 0 if x < 220 else 1, '1')

                #if avgBkg <= (self._Bkg / 100.0):
                #    tile.save(outfile, quality=self._quality)
                tile.save(outfile, quality=self._quality)
                self._queue.task_done()
            except Exception as e:
                # print(level, address)
                print("image %s failed at dz.get_tile for level %f" %
                      (self._slidepath, level))
                # e = sys.exc_info()[0]
                print(e)
                self._queue.task_done()

    def _get_dz(self, associated=None):

        image = self._slide

        return DeepZoomGenerator(image, self._tile_size, self._overlap, limit_bounds=self._limit_bounds)


class DeepZoomImageTiler(object):
    """Handles generation of tiles and metadata for a single image."""

    def __init__(self, dz, basename, labelimage_path, tile_size, format, queue, slide, basenameJPG, ImgExtension, Mag):
        self._dz = dz
        self._basename = basename
        self._labelimage_path = labelimage_path
        self._tile_size = tile_size
        self._basenameJPG = basenameJPG
        self._format = format
        self._queue = queue
        self._processed = 0
        self._slide = slide
        self._ImgExtension = ImgExtension
        self._Mag = Mag

    def run(self):
        self._write_tiles()

    def _write_tiles(self):
        Magnification = 20
 
        Factors = self._slide.level_downsamples #每个 level 下采样的倍数
        print("svs.level_downsamples: ", Factors) 
        print("svs.level_count: ", self._slide.level_count)
        
        
        try:
            Objective = float(
                self._slide.properties[openslide.PROPERTY_NAME_OBJECTIVE_POWER])
        except:
            print(self._basename + " - No Obj information found")
            print(self._ImgExtension)
            if ("jpg" in self._ImgExtension) | ("dcm" in self._ImgExtension) | ("tif" in self._ImgExtension):
                # Objective = self._ROIpc
                Objective = 1.
                Magnification = Objective
                print("input is jpg - will be tiled as such with %f" % Objective)
            else:
                return
        
        print("Objective: ", Objective)
        Available = tuple(Objective / x for x in Factors)
        print("Available: ", Available)
        Mismatch = tuple(x-Magnification for x in Available)
        AbsMismatch = tuple(abs(x) for x in Mismatch)
        if len(AbsMismatch) < 1:
            print(self._basename + " - Objective field empty!")
            return

        for level in range(self._dz.level_count-1, -1, -1):
            ThisMag = Available[0]/pow(2, self._dz.level_count-(level+1)) # 20.0
            
            if self._Mag > 0:
                if ThisMag != self._Mag:
                    continue
                    
            print("level: ", level)
            print("ThisMag: ", ThisMag)
            tiledir = os.path.join("%s" % self._basename, str(ThisMag))
            
            if not os.path.exists(self._labelimage_path + '/' + self._basenameJPG + '-labels.png'):
                continue
                
            if not os.path.exists(tiledir):
                os.makedirs(tiledir)

            cols, rows = self._dz.level_tiles[level]
            print("cols, rows", cols, rows)
            

            
            fp = open(self._labelimage_path + '/' + self._basenameJPG + '-labels.png', 'rb')
            label_image = Image.open(fp).convert('L').point(lambda x: 1 if x < 220 else 0, 'F')
            label_array = np.array(label_image)
            fp.close()
            (max_row, max_col) = label_array.shape
            print("labelimage size: ", max_row, max_col)
            
            label_multiple = round((self._tile_size / label_downsample) * (Objective / ThisMag))
            print("labelimage multiple: ", label_multiple)
            
            print(max_col / label_multiple) 
            print(max_row / label_multiple)

            for row in range(rows):
                for col in range(cols):
                    cur_label_array_rows = label_array[row * label_multiple : min(max_row, (row + 1) * label_multiple), :]
                    cur_label_array = cur_label_array_rows[: , col * label_multiple : min(max_col, (col + 1) * label_multiple)]
                    #print(row * label_multiple)
                    #print(min(max_row, (row + 1) * label_multiple))
                    #print(col * label_multiple )
                    #print(min(max_col, (col + 1) * label_multiple))
                    #print(cur_label_array)
                    avglabel = np.average(cur_label_array)
                    #print(avglabel)
                    
                    
                    patch_label = 1 if avglabel > tumor_threshold else 0
                    if patch_label == 0:
                        continue
                    tilename = os.path.join( tiledir,
                        '%d_%d_%d.%s' % (col, row, patch_label, self._format))  # add a label, 1 Tomor, 0 Normal
                    #tilename = os.path.join( tiledir,
                    #    '%d_%d_%.2f.%s' % (col, row, avglabel, self._format))  # add a label

                    if not os.path.exists(tilename):
                        self._queue.put((level, (col, row), tilename))
                    self._tile_done()

    def _tile_done(self):
        self._processed += 1
        count, total = self._processed, self._dz.tile_count
        
        print("Tiling %s: wrote %d/%d tiles" % 
              ('slide', count, total), end='\r', file=sys.stderr)
        if count == total:
            print(file=sys.stderr)


class DeepZoomStaticTiler(object):
    """Handles generation of tiles and metadata for all images in a slide."""

    def __init__(self, slidepath, basename, labelimage_path, format, tile_size, overlap,
                limit_bounds, quality, workers, Bkg, basenameJPG, ImgExtension, Mag):

        self._slide = open_slide(slidepath)
        self._basename = basename
        self._labelimage_path = labelimage_path
        self._basenameJPG = basenameJPG
        self._format = format
        self._tile_size = tile_size
        self._overlap = overlap
        self._limit_bounds = limit_bounds
        self._queue = JoinableQueue(2 * workers)
        self._workers = workers
        self._Bkg = Bkg
        self._dzi_data = {}
        self._ImgExtension = ImgExtension
        self._Mag = Mag
        

        for _i in range(workers):
            TileWorker(self._queue, slidepath, tile_size, overlap,
                limit_bounds, quality, self._Bkg).start()

    def run(self):
        self._run_image()
        self._shutdown()

    def _run_image(self):
        """Run a single image from self._slide."""
        image = self._slide
        basename = self._basename

        dz = DeepZoomGenerator(image, self._tile_size,
                               self._overlap, limit_bounds=self._limit_bounds)
        
        print("dz.level_tiles", dz.level_tiles)
        print("dz.level_dimensions", dz.level_dimensions)

        tiler = DeepZoomImageTiler(
            dz,
            basename,
            self._labelimage_path,
            self._tile_size,
            self._format,
            self._queue,
            self._slide,
            self._basenameJPG,
            self._ImgExtension,
            self._Mag)

        tiler.run()

    def _shutdown(self):
        for _i in range(self._workers):
            self._queue.put(None)
        self._queue.join()


if __name__ == '__main__':
    parser = OptionParser(usage='Usage: %prog [options] <slide>')
    # python 0_generate_patch.py -j 16 -o '/data15/data15_5/Public/Datasets/stomach/stomach_patch_20x_cancer' -l '/data15/data15_5/Public/Datasets/lj_stomach/innerlabel' -s 512  -e 0 -B 50 -M 20 "/data15/data15_5/Public/Datasets/lj_stomach/data/*/*.svs"
    parser.add_option('-L', '--ignore-bounds', 
        dest='limit_bounds', default=True, action='store_false', help='display entire scan area')
    parser.add_option('-e', '--overlap', metavar='PIXELS', 
        dest='overlap', type='int', default=1, help='overlap of adjacent tiles [1]')
    parser.add_option('-f', '--format', metavar='{jpeg|png}', 
        dest='format', default='jpeg', help='image format for tiles [jpeg]')
    parser.add_option('-j', '--jobs', metavar='COUNT', 
        dest='workers', type='int', default=4, help='number of worker processes to start [4]')
    parser.add_option('-o', '--output', metavar='NAME', 
        dest='output_path', help='base name of output file')
    parser.add_option('-l', '--labelimage', metavar='NAME', 
        dest='labelimage_path', help='base name of label image file')
    parser.add_option('-Q', '--quality', metavar='QUALITY', 
        dest='quality', type='int', default=90, help='JPEG compression quality [90]')
    parser.add_option('-s', '--size', metavar='PIXELS', 
        dest='tile_size', type='int', default=254, help='tile size [254]')
    parser.add_option('-B', '--Background', metavar='PIXELS', 
        dest='Bkg', type='float', default=50, help='Max background threshold [50]; percentager of background allowed')
    parser.add_option('-M', '--Mag', metavar='PIXELS', 
        dest='Mag', type='float', default=-1, help='Magnification at which tiling should be done (-1 of all)')

    (opts, args) = parser.parse_args()

    try:
       slidepath = args[0]
    except IndexError:
       parser.error('Missing slide argument')

    files = glob(slidepath)  # get all image paths from data folder
    ImgExtension = slidepath.split('*')[-1]  # eg. svs

    print('Read images from :', slidepath)
    print('Image number :', len(files))
    print('Extension :', ImgExtension)
    print("***********************\n"*3)
    
    #xlsxFile = '/data16/Public/Datasets/TCGA-STAD/label.xlsx'
    #label_dict = load_xlsx_file(xlsxFile)

    files = sorted(files)
    
    for imgNb, filename in enumerate(files):
        basename = os.path.splitext(os.path.basename(filename))[0]   # get file basename
        # if label_dict[basename] == 'Normal':
        #    continue

        result_folder_path = os.path.join(opts.output_path, basename)
        print('\n\nload file: ', filename)
        print("processing: " + basename + " with extension: " + ImgExtension)
        print('save results to: ', result_folder_path)

        if os.path.exists(result_folder_path):
            print("Image %s already tiled" % basename)
            continue
        try:
            DeepZoomStaticTiler(    
                filename, 
                result_folder_path,
                opts.labelimage_path, 
                opts.format, 
                opts.tile_size, 
                opts.overlap, 
                opts.limit_bounds, 
                opts.quality, 
                opts.workers, 
                opts.Bkg, 
                basename, 
                ImgExtension, 
                opts.Mag).run()
        except Exception as e:
            print("Failed to process file %s, error: %s" % (filename, sys.exc_info()[0]))
            print(e)

    print("End")
